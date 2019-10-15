from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import BLACK
from collections import namedtuple
from config.config import DATA_PATH
from common.RecordLog import log
from common.DataReplace import DataReplace


class ParseExcel(object):
    """解析excel文件"""
    def __init__(self,filename):
        try:
            self.filename = filename
            self.__wb = load_workbook(self.filename)
        except FileNotFoundError as e :
            log.info("解析Excel文件{}失败\n{}".format(self.filename, e))

    def get_max_row_num(self,sheet_name):
        """获取最大行号"""
        max_row_num = self.__wb[sheet_name].max_row
        return max_row_num

    def get_max_column_num(self,sheet_name):
        """获取最大的列号"""
        max_column = self.__wb[sheet_name].max_column
        return max_column

    # coordinate 为单元格的格式列如：A5
    def get_cell_value(self,sheet_name,coordinate=None,row=None,column=None):
        """获取指定单元格的数据"""
        if coordinate is not None:
            try:
                return self.__wb[sheet_name][coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row,int) and isinstance(column,int):
                return self.__wb[sheet_name].cell(row=row,column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self,sheet_name,row):
        """获取某一行的数据"""
        column_num = self.get_max_column_num(sheet_name)
        row_value = []
        if isinstance(row,int):
            for column in range(1,column_num+1):
                values_row = self.__wb[sheet_name].cell(row,column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError("row must be type int")

    def get_column_value(self,sheet_name,column):
        """获取某一列数据"""
        row_num = self.get_max_row_num(sheet_name,column)
        column_value = []
        if isinstance(column,int):
            for row in range(1,row_num+1):
                value_column = self.__wb[sheet_name].cell(row,column).value
                column_value.append(value_column)
            return column_value
        else:
            raise TypeError("row must be type int")


    def get_all_value(self,sheet_name):
        """获取指定表单的所有数据（除去表头）"""
        rows_obj = self.__wb[sheet_name].iter_rows(min_row=2,max_row=self.__wb[sheet_name].max_row,values_only=True)
        values = []
        for row_tuple in rows_obj:
            values_list = []
            for value in row_tuple:
                values_list.append(value)
            values.append(values_list)
        log.info("读取{}文件，表单{}的所有数据\n{}".format(self.filename, sheet_name, values))
        return values


    def get_excel_title(self, sheet_name):
        """获取sheet表头"""
        title_key = tuple(self.__wb[sheet_name].iter_rows(max_row=1, values_only=True))[0]
        log.info("解析{}文件表单{}的标题:\n{}".format(self.filename, sheet_name, title_key))
        return title_key

    def get_list_dict_all_value(self, sheet_name):
        """获取所有数据，返回嵌套字典的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        all_values = self.get_all_value(sheet_name)
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(sheet_title, value)))
        return value_list

    def get_name_tuple_all_value(self, sheet_name):
        """获取所有数据，返回嵌套命名元组的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        values = self.get_all_value(sheet_name)
        excel = namedtuple('excel', sheet_title)
        value_list = []
        for value in values:
            e = excel(*value)
            value_list.append(e)
        return value_list


    def write_cell(self, sheet_name, row, column, value=None, bold=True, color=BLACK):
        if isinstance(row, int) and isinstance(column, int):
            try:
                log.info("{}文件，表单{},第{}行第{}列写入数据{}".format(self.filename, sheet_name, row, column, value))
                cell_obj = self.__wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=color, bold=bold)
                cell_obj.value = value
                self.__wb.save(self.filename)
            except Exception as e:
                log.error("{}文件，表单{},第{}行第{}列写入数据{}失败\n{}".
                          format(self.filename, sheet_name, row, column, value, e))
                raise e
        else:
            log.error("{}文件写数据失败：row and column must be type int".format(self.filename))
            raise TypeError('row and column must be type int')


do_excel = ParseExcel(DATA_PATH)

if __name__ == "__main__":
    import re

    # dc = ParseExcel(DATA_PATH)
    # data = dc.get_cell_value("test","A1")
    # print(data)
    # num = "300"
    # add = "\$\{add}"
    # b = DataReplace.re_replace(add,num,data)
    # print(b)
    dc = ParseExcel(DATA_PATH)
    pc = dc.get_name_tuple_all_value("class")[0]
    data = dc.get_name_tuple_all_value("class")[0][3]
    print(data)
    import json
    bc = json.loads(data)
    ac = bc["vcode"]
    print(ac)




